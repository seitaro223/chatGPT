#!/usr/bin/env python3
# BT10 F_form linear_v1 independent reimplementation + full-set diff audit
#
# linear_v1 definition (reimplemented from scratch; legacy MATLAB code NOT reused):
#   x_DNB       = DNB_position / heated_length
#   f_DNB       = interp1(x, f, x_DNB)                  # linear interpolation
#   Blue_area   = integral_0^x_DNB f(x) dx              # trapezoid + partial last segment
#   Orange_area = x_DNB * f_DNB
#   F_form      = Blue_area / Orange_area
#
# Source of truth for f(x): バンドルデータ整理r3.xlsx, sheets
#   '非一様加熱を一様加熱に補正108/161/164', columns E (z/L) and F (f).
# Recorded values reconciled against:
#   BT08A3_macro_Fform_replace_map_20260616_145859.csv  (116 rows)
#   H52Q_current_bundle_input_v2b_FformLinear_20260616_145435.xlsx
#
# Run from the 20260615_MATLAB directory.

import openpyxl

R3 = "バンドルデータ整理r3.xlsx"

# master case -> (profile sheet, n profile rows, x_DNB, recorded linear, recorded legacy)
CASES = [
    ("108_70in",         "非一様加熱を一様加熱に補正108", 28, 0.729167, 0.6198206589, 0.654327947937632),
    ("108_76in",         "非一様加熱を一様加熱に補正108", 28, 0.791667, 0.7336799376, 0.760494),
    ("161_uniform",      "非一様加熱を一様加熱に補正161 ", 28, 0.997024, 1.0,          1.0),
    ("164_112in",        "非一様加熱を一様加熱に補正164", 27, 0.666667, 0.8776442967, 1.014),
    ("164_134in_normal", "非一様加熱を一様加熱に補正164", 27, 0.797619, 1.2999931692, 1.363),
]


def get_profile(wb, sheet, nrows):
    ws = wb[sheet]
    x, f = [], []
    for r in range(2, 2 + nrows):
        x.append(float(ws.cell(r, 5).value))  # col E = z/L
        f.append(float(ws.cell(r, 6).value))  # col F = f(x)
    return x, f


def interp(x, f, xq):
    if xq <= x[0]:
        return f[0]
    if xq >= x[-1]:
        return f[-1]
    for i in range(len(x) - 1):
        if x[i] <= xq <= x[i + 1]:
            t = (xq - x[i]) / (x[i + 1] - x[i])
            return f[i] + t * (f[i + 1] - f[i])
    return None


def blue_area(x, f, xD):
    """Integral 0..xD, partial last segment by linear interpolation (clean linear_v1)."""
    a = 0.0
    for i in range(len(x) - 1):
        x0, x1 = x[i], x[i + 1]
        if x1 <= xD:
            a += 0.5 * (f[i] + f[i + 1]) * (x1 - x0)
        elif x0 < xD < x1:
            fq = interp(x, f, xD)
            a += 0.5 * (f[i] + fq) * (xD - x0)
            break
        else:
            break
    return a


def blue_inclusive(x, f, xD):
    """SUM that runs one grid node PAST xD (downstream-of-DNB contamination)."""
    a = 0.0
    for i in range(len(x) - 1):
        x0, x1 = x[i], x[i + 1]
        if x1 <= xD:
            a += 0.5 * (f[i] + f[i + 1]) * (x1 - x0)
        elif x0 < xD < x1:
            a += 0.5 * (f[i] + f[i + 1]) * (x1 - x0)
            break
        else:
            break
    return a


def main():
    wb = openpyxl.load_workbook(R3, data_only=True)
    print("case | source_file | F_form_recomputed | recorded_linear | diff | ratio | flag")
    for lab, sh, nr, xD, rlin, rleg in CASES:
        x, f = get_profile(wb, sh, nr)
        fD = interp(x, f, xD)
        F = blue_area(x, f, xD) / (xD * fD)
        d = F - rlin
        ratio = F / rlin if rlin else float("nan")
        flag = "FLAG_gt_0.005" if abs(d) > 0.005 else "ok"
        print(f"{lab} | {R3}:{sh} | {F:.6f} | {rlin:.6f} | {d:+.6f} | {ratio:.6f} | {flag}")

    print()
    print("case | F_form_recomputed | recorded_legacy | diff | ratio | flag")
    for lab, sh, nr, xD, rlin, rleg in CASES:
        x, f = get_profile(wb, sh, nr)
        fD = interp(x, f, xD)
        F = blue_area(x, f, xD) / (xD * fD)
        d = F - rleg
        ratio = F / rleg if rleg else float("nan")
        flag = "FLAG_gt_0.005" if abs(d) > 0.005 else "ok"
        print(f"{lab} | {F:.6f} | {rleg:.6f} | {d:+.6f} | {ratio:.6f} | {flag}")

    print()
    print("=== contamination decomposition (legacy blue vs clean linear blue) ===")
    print("case | blue_clean | blue_one_node_past | legacy_blue_used | note")
    legacy_blue = {
        "108_70in": 0.755174, "108_76in": None, "161_uniform": 1.0,
        "164_112in": 0.8596465, "164_134in_normal": 0.926529,
    }
    for lab, sh, nr, xD, rlin, rleg in CASES:
        x, f = get_profile(wb, sh, nr)
        bc = blue_area(x, f, xD)
        bi = blue_inclusive(x, f, xD)
        lb = legacy_blue[lab]
        print(f"{lab} | {bc:.6f} | {bi:.6f} | {lb} ")


if __name__ == "__main__":
    main()
