#!/usr/bin/env python3
"""Statically inject TM03C main280 inputs into a B2 robust xlsm.

No Excel, no MATLAB, no Solver, no new VBA modules.
Uses openpyxl keep_vba=True and preserves the existing VBA project.
"""
from __future__ import annotations

import copy
import datetime as dt
import json
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter

ROOT = Path(__file__).resolve().parent
PREFERRED_TEMPLATE = ROOT / "TM03Cprep_B2robust_B1b12_injected_20260626_113807.xlsm"
FALLBACK_TEMPLATE = ROOT / "TM03B2_macro_bracket_test_20260625_235759_run_done.xlsm"
PREFERRED_INPUT = ROOT / "TM03C_main280_input_20260626_130059.xlsx"
FALLBACK_PREVIEW = ROOT / "tm_input_preview_20260625_083115.xlsx"

START_ROW = 226
END_ROW = 505
EXPECTED_N = 280
FORMULA_TEMPLATE_ROW = 88
LAST_COL = "BR"
TABLE_NAME = "テーブル2"
TM_SHEET = "tm"

INPUT_COLUMNS = {
    "A": "No_TableNo",
    "B": "P",
    "M": "No",
    "N": "G",
    "Q": "DH",
    "R": "L_DNB",
    "S": "q_in",
    "T": "Tin",
    "V": "f",
    "X": "f_balance_seed",
    "AC": "Tw",
    "AG": "y_star",
    "AP": "UB",
    "BE": "q_M",
    "BG": "F_form",
    "BH": "x_Mes",
    "BI": "A_corr",
    "BJ": "sigma_corr",
    "BK": "Fcorr",
    "BQ": "F2",
    "BR": "L",
}

FIXED_VALUES = {
    "V": 0.03,
    "X": 0.01,
    "AC": 600,
    "AG": 1e-5,
    "AP": 1,
    "BG": 1,
    "BI": 0.046,
    "BJ": 5625,
    "BK": 1,
    "BQ": 1,
}

PREVIEW_ALIASES = {
    "No_TableNo": ["tmA_No_TableNo", "A_No_TableNo", "No_TableNo"],
    "P": ["tmB_P_Pa", "B_P", "P", "P_Pa"],
    "No": ["M_No", "ExptNo", "No"],
    "G": ["tmN_G_kg_m2s", "N_G", "G", "G_kg_m2s"],
    "DH": ["tmQ_DH_m", "Q_DH", "DH", "D_H", "Dh"],
    "L_DNB": ["tmR_L_DNB_m", "R_L_DNB", "L_DNB", "LDNB", "z_DNB"],
    "q_M": ["tmBE_qM_W_m2", "BE_q_M", "q_M", "qM", "q_exp", "S_q_in_seed"],
    "Tin": ["tmT_Tin_K", "T_Tin", "Tin", "Tin_K"],
    "x_Mes": ["tmBH_x_Mes", "BH_x_Mes", "x_Mes", "x_report", "x_eq"],
    "L": ["tmBR_L_m", "BR_L", "L", "Length"],
    "key": ["key", "No_TableNo", "tmA_No_TableNo", "A_No_TableNo"],
    "TableNo": ["TableNo", "Table"],
}

INPUT_ALIASES = {
    "No_TableNo": ["A_No_TableNo", "No_TableNo"],
    "P": ["B_P", "P"],
    "No": ["M_No", "No", "ExptNo"],
    "G": ["N_G", "G"],
    "DH": ["Q_DH", "DH"],
    "L_DNB": ["R_L_DNB", "L_DNB"],
    "q_M": ["BE_q_M", "q_M", "S_q_in_seed"],
    "Tin": ["T_Tin", "Tin"],
    "x_Mes": ["BH_x_Mes", "x_Mes"],
    "L": ["BR_L", "L"],
    "key": ["key", "No_TableNo", "A_No_TableNo"],
    "TableNo": ["TableNo", "Table"],
}


def norm(name: str) -> str:
    return "".join(ch.lower() for ch in str(name) if ch.isalnum())


def has_vba(path: Path) -> bool:
    with zipfile.ZipFile(path, "r") as zf:
        return "xl/vbaProject.bin" in zf.namelist()


def read_sheet_rows(path: Path, sheet_name: str | None = None) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v) if v is not None else "" for v in rows[0]]
    out = []
    for vals in rows[1:]:
        if all(v is None for v in vals):
            continue
        out.append({headers[i]: vals[i] if i < len(vals) else None for i in range(len(headers))})
    wb.close()
    return out


def pick(row: dict, logical_name: str, aliases: dict[str, list[str]]):
    by_norm = {norm(k): v for k, v in row.items()}
    for candidate in aliases[logical_name]:
        key = norm(candidate)
        if key in by_norm:
            return by_norm[key]
    raise KeyError(f"Missing {logical_name}; aliases={aliases[logical_name]}")


def load_main280() -> tuple[list[dict], Path, list[str]]:
    warnings: list[str] = []
    if PREFERRED_INPUT.exists():
        source = PREFERRED_INPUT
        raw = read_sheet_rows(source, "input")
        aliases = INPUT_ALIASES
    else:
        source = FALLBACK_PREVIEW
        warnings.append(f"Preferred input xlsx not found: {PREFERRED_INPUT.name}; used {source.name} instead.")
        raw = read_sheet_rows(source, "tm_input_preview")
        aliases = PREVIEW_ALIASES
    if len(raw) != EXPECTED_N:
        raise RuntimeError(f"Expected {EXPECTED_N} input rows, found {len(raw)} in {source}")
    data = []
    for idx, row in enumerate(raw, START_ROW):
        q_m = pick(row, "q_M", aliases)
        rec = {
            "target_row": idx,
            "No_TableNo": pick(row, "No_TableNo", aliases),
            "P": pick(row, "P", aliases),
            "No": pick(row, "No", aliases),
            "G": pick(row, "G", aliases),
            "DH": pick(row, "DH", aliases),
            "L_DNB": pick(row, "L_DNB", aliases),
            "q_in": q_m,
            "Tin": pick(row, "Tin", aliases),
            "q_M": q_m,
            "x_Mes": pick(row, "x_Mes", aliases),
            "L": pick(row, "L", aliases),
            "key": pick(row, "key", aliases),
            "TableNo": pick(row, "TableNo", aliases),
        }
        for col, value in FIXED_VALUES.items():
            rec[INPUT_COLUMNS[col]] = value
        data.append(rec)
    return data, source, warnings


def copy_row_template(ws, src_row: int, dst_row: int) -> None:
    for col_idx in range(1, column_index_from_string(LAST_COL) + 1):
        src = ws.cell(src_row, col_idx)
        dst = ws.cell(dst_row, col_idx)
        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
        else:
            dst.value = src.value
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy.copy(src.font)
        if src.fill:
            dst.fill = copy.copy(src.fill)
        if src.border:
            dst.border = copy.copy(src.border)
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.protection:
            dst.protection = copy.copy(src.protection)
        if src.comment:
            dst.comment = copy.copy(src.comment)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def inject_inputs(ws, data: list[dict]) -> None:
    for rec in data:
        r = rec["target_row"]
        for col, field in INPUT_COLUMNS.items():
            ws[f"{col}{r}"] = rec[field]


def resize_table(ws) -> tuple[bool, str]:
    if TABLE_NAME not in ws.tables:
        return False, f"{TABLE_NAME} not found; existing={list(ws.tables.keys())}"
    table = ws.tables[TABLE_NAME]
    table.ref = f"A1:{LAST_COL}{END_ROW}"
    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref
    return True, table.ref


def write_qa(path: Path, data: list[dict], qa_rows: list[dict], colmap: list[dict], report_info: dict) -> None:
    wb = Workbook()
    default = wb.active
    default.title = "QA_checks"

    def write_table(ws, rows: list[dict]):
        if not rows:
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])

    write_table(default, qa_rows)
    ws = wb.create_sheet("injected_points")
    write_table(ws, [{"target_row": r["target_row"], "No_TableNo": r["No_TableNo"], "key": r["key"], "TableNo": r["TableNo"]} for r in data])
    ws = wb.create_sheet("input_values")
    value_fields = ["target_row", "No_TableNo", "P", "No", "G", "DH", "L_DNB", "q_in", "Tin", "q_M", "x_Mes", "L"]
    write_table(ws, [{k: r[k] for k in value_fields} for r in data])
    ws = wb.create_sheet("column_mapping")
    write_table(ws, colmap)
    ws = wb.create_sheet("table_info")
    write_table(ws, [report_info])
    ws = wb.create_sheet("batch_plan")
    batches = []
    start = START_ROW
    batch = 1
    while start <= END_ROW:
        end = min(start + 49, END_ROW)
        batches.append({"batch": batch, "start_row": start, "end_row": end, "N": end - start + 1})
        start = end + 1
        batch += 1
    write_table(ws, batches)
    wb.save(path)


def main() -> None:
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    template = PREFERRED_TEMPLATE if PREFERRED_TEMPLATE.exists() else FALLBACK_TEMPLATE
    if not template.exists():
        raise FileNotFoundError(f"No template xlsm found: {PREFERRED_TEMPLATE} or {FALLBACK_TEMPLATE}")
    data, input_source, warnings = load_main280()

    out_xlsm = ROOT / f"TM03C_B2robust_main280_injected_by_codex_{stamp}.xlsm"
    qa_xlsx = ROOT / f"TM03C_B2robust_main280_injected_by_codex_QA_{stamp}.xlsx"
    report_md = ROOT / f"run_report_TM03C_B2robust_main280_injected_by_codex_{stamp}.md"

    before_vba = has_vba(template)
    wb = load_workbook(template, keep_vba=True, data_only=False)
    ws = wb[TM_SHEET]
    for r in range(START_ROW, END_ROW + 1):
        copy_row_template(ws, FORMULA_TEMPLATE_ROW, r)
    inject_inputs(ws, data)
    table_ok, table_msg = resize_table(ws)
    wb.save(out_xlsm)
    wb.close()
    after_vba = has_vba(out_xlsm)

    # Re-open for QA against saved file.
    qwb = load_workbook(out_xlsm, keep_vba=True, data_only=False)
    qws = qwb[TM_SHEET]
    actual_ids = [qws[f"A{r}"].value for r in range(START_ROW, END_ROW + 1)]
    expected_ids = [r["No_TableNo"] for r in data]
    duplicate_count = len(expected_ids) - len(set(expected_ids))
    fixed_ok = all(
        qws[f"{col}{r['target_row']}"].value == value
        for r in data
        for col, value in FIXED_VALUES.items()
    )
    s_be_ok = all(
        qws[f"S{r['target_row']}"].value == qws[f"BE{r['target_row']}"].value == r["q_M"]
        for r in data
    )
    inputs_ok = all(
        qws[f"{col}{r['target_row']}"].value == r[field]
        for r in data
        for col, field in INPUT_COLUMNS.items()
    )
    formula_cells_nonblank = sum(
        1
        for row in qws.iter_rows(min_row=START_ROW, max_row=END_ROW, min_col=1, max_col=column_index_from_string(LAST_COL))
        for cell in row
        if cell.value is not None
    )
    table_ref = qws.tables[TABLE_NAME].ref if TABLE_NAME in qws.tables else "missing"
    qwb.close()

    qa_rows = [
        {"check": "280_points_injected", "pass": len(actual_ids) == EXPECTED_N},
        {"check": "No_TableNo_order_matches_input", "pass": actual_ids == expected_ids},
        {"check": "No_TableNo_no_duplicates", "pass": duplicate_count == 0},
        {"check": "rows_226_to_505", "pass": data[0]["target_row"] == START_ROW and data[-1]["target_row"] == END_ROW},
        {"check": "A1_BR505_table_range", "pass": table_ref == f"A1:{LAST_COL}{END_ROW}"},
        {"check": "input_21_columns_match_spec", "pass": inputs_ok},
        {"check": "S_equals_BE_equals_q_M", "pass": s_be_ok},
        {"check": "fixed_values_match_spec", "pass": fixed_ok},
        {"check": "vba_project_preserved", "pass": before_vba and after_vba},
        {"check": "template_row_formulas_values_copied_nonblank_count", "pass": formula_cells_nonblank > 0, "detail": formula_cells_nonblank},
    ]
    for w in warnings:
        qa_rows.append({"check": "warning", "pass": False, "detail": w})
    if not table_ok:
        qa_rows.append({"check": "table_resize_warning", "pass": False, "detail": table_msg})

    colmap = [{"tm_column": col, "field": field, "source": "fixed" if col in FIXED_VALUES else "main280_input"} for col, field in INPUT_COLUMNS.items()]
    info = {
        "template": template.name,
        "input_source": input_source.name,
        "output_xlsm": out_xlsm.name,
        "qa_xlsx": qa_xlsx.name,
        "table_ref": table_ref,
        "vba_before": before_vba,
        "vba_after": after_vba,
        "warnings": json.dumps(warnings, ensure_ascii=False),
    }
    write_qa(qa_xlsx, data, qa_rows, colmap, info)

    report = [
        "# TM03C B2 robust main280 injected by Codex",
        "",
        "## Purpose",
        "Statically inject main280 rows 226-505 into the B2 robust xlsm without Excel, MATLAB, Solver, or new VBA.",
        "",
        "## Inputs",
        f"- Template xlsm: `{template.name}`",
        f"- main280 source: `{input_source.name}`",
        "",
        "## Output",
        f"- Injected xlsm: `{out_xlsm.name}`",
        f"- QA workbook: `{qa_xlsx.name}`",
        "",
        "## Method",
        f"- Loaded xlsm with `keep_vba=True`; VBA project present before save: `{before_vba}`; after save: `{after_vba}`.",
        f"- Copied `tm!A{FORMULA_TEMPLATE_ROW}:{LAST_COL}{FORMULA_TEMPLATE_ROW}` to rows {START_ROW}-{END_ROW}; formulas were translated row-relative with openpyxl Translator.",
        "- Overwrote the 21 specified input columns after formula/style copy.",
        f"- Resized `{TABLE_NAME}` to `A1:{LAST_COL}{END_ROW}` where possible; actual: `{table_ref}`.",
        "",
        "## QA summary",
    ]
    for row in qa_rows:
        report.append(f"- {row['check']}: {row['pass']}" + (f" ({row.get('detail')})" if row.get("detail") else ""))
    report += [
        "",
        "## Next step",
        "Open the generated xlsm in Windows Excel, enable macros and Solver, then run the existing B2 robust macro for Batch 1 only: start row 226, end row 275.",
    ]
    report_md.write_text("\n".join(report) + "\n", encoding="utf-8")
    print(out_xlsm)
    print(qa_xlsx)
    print(report_md)


if __name__ == "__main__":
    main()
