#!/usr/bin/env python3
"""Build the TM03I3 source09 appended workbook copy locally.

This script intentionally does not run Excel, VBA macros, or recalculation.
It copies the r129 source workbook, appends the 117 TM03I2 preview rows to
`tm`, copies formulas from row 502, extends table `テーブル2` to A1:BR619,
and writes the copied xlsm next to the source workbook.

The generated xlsm is not committed because binary workbook artifacts can block
PR creation/review. Run this locally when the workbook artifact is needed.
"""
from __future__ import annotations

import copy
import csv
import os
import shutil
from pathlib import Path

import openpyxl
from iapws import IAPWS97
from openpyxl.formula.translate import Translator

BASE = Path(__file__).resolve().parent
SOURCE_XLSM = BASE / "celataモデル_簡易計算_単管_櫻井検算r129_F1なし_09追加予定.xlsm"
PREVIEW_CSV = BASE / "TM03I2_source09_append_rows_preview.csv"
OUTPUT_XLSM = BASE / "celataモデル_簡易計算_単管_櫻井検算r129_F1なし_09追加予定_TM03I3_source09追記済みコピー.xlsm"

VALUE_COLUMNS = [
    "No_TableNo",
    "P",
    "No",
    "G",
    "DH",
    "L_DNB",
    "q_in",
    "Tin",
    "q_M",
    "F_form",
    "x_Mes",
    "Fcorr",
    "F2",
    "L",
]


def as_number(value: str) -> float:
    return float(value)


def tin_from_hsub(p_pa: float, hsub_kjkg: float) -> float:
    """Return Tin [K] using IF97: h_in = h_f(P) - Hsub."""
    p_mpa = p_pa / 1e6
    hf = IAPWS97(P=p_mpa, x=0).h
    hin = hf - hsub_kjkg
    return IAPWS97(P=p_mpa, h=hin).T


def main() -> None:
    if not SOURCE_XLSM.exists():
        raise FileNotFoundError(SOURCE_XLSM)
    if not PREVIEW_CSV.exists():
        raise FileNotFoundError(PREVIEW_CSV)

    if OUTPUT_XLSM.exists():
        OUTPUT_XLSM.unlink()
    shutil.copy2(SOURCE_XLSM, OUTPUT_XLSM)

    wb = openpyxl.load_workbook(OUTPUT_XLSM, keep_vba=True, data_only=False)
    ws = wb["tm"]
    table = ws.tables["テーブル2"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_to_col = {h: i + 1 for i, h in enumerate(headers)}

    if ws.max_row != 502 or ws.max_column != 70 or table.ref != "A1:BR502":
        raise RuntimeError(
            f"Unexpected tm structure: row={ws.max_row}, col={ws.max_column}, table={table.ref}"
        )

    with PREVIEW_CSV.open(newline="", encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    if len(rows) != 117:
        raise RuntimeError(f"Expected 117 preview rows, got {len(rows)}")

    template_row = 502
    start_row = 503
    for offset, srcrow in enumerate(rows):
        row_no = start_row + offset
        for col_no in range(1, 71):
            src_cell = ws.cell(template_row, col_no)
            dst_cell = ws.cell(row_no, col_no)
            if src_cell.data_type == "f" and isinstance(src_cell.value, str):
                dst_cell.value = Translator(
                    src_cell.value, origin=src_cell.coordinate
                ).translate_formula(dst_cell.coordinate)
            else:
                dst_cell.value = src_cell.value

            if src_cell.has_style:
                dst_cell._style = copy.copy(src_cell._style)
            dst_cell.number_format = src_cell.number_format
            dst_cell.font = copy.copy(src_cell.font)
            dst_cell.fill = copy.copy(src_cell.fill)
            dst_cell.border = copy.copy(src_cell.border)
            dst_cell.alignment = copy.copy(src_cell.alignment)
            dst_cell.protection = copy.copy(src_cell.protection)

        for header in VALUE_COLUMNS:
            col_no = header_to_col[header]
            if header == "Tin":
                value = tin_from_hsub(
                    as_number(srcrow["P"]), as_number(srcrow["_Hsub_kJkg"])
                )
            else:
                value = srcrow.get(header, "")
            if value == "":
                ws.cell(row_no, col_no).value = None
            elif header == "No_TableNo":
                ws.cell(row_no, col_no).value = value
            else:
                ws.cell(row_no, col_no).value = as_number(value)

    table.ref = "A1:BR619"
    wb.save(OUTPUT_XLSM)

    check_wb = openpyxl.load_workbook(OUTPUT_XLSM, keep_vba=True, data_only=False)
    check_ws = check_wb["tm"]
    print(f"wrote: {OUTPUT_XLSM}")
    print(f"tm rows={check_ws.max_row} cols={check_ws.max_column}")
    print(f"テーブル2={check_ws.tables['テーブル2'].ref}")
    print(f"first_id={check_ws['A503'].value} last_id={check_ws['A619'].value}")


if __name__ == "__main__":
    main()
